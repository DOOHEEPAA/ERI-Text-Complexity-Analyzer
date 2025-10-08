import pandas as pd
from kiwipiepy import Kiwi # 한국어 형태소 분석기 라이브러리
import re
import os
import tkinter as tk # GUI 생성을 위한 라이브러리
from tkinter import messagebox
from openpyxl import load_workbook # 엑셀 파일 스타일링을 위한 라이브러리
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from typing import Set, Tuple, List, Dict # 타입 힌팅을 위한 모듈

# ==============================================================================
# 1. 설정 클래스 (Configuration)
# ==============================================================================
class ERIConfig:
    """ERI(Explainable Readability Index) 계산에 필요한 모든 설정을 관리하는 클래스"""
    
    # --- 파일 경로 설정 ---
    # __file__이 정의되어 있으면 스크립트 파일 기준, 아니면 현재 작업 디렉토리 기준으로 설정
    BASE_DIR = os.path.dirname(os.path.abspath(__file__)) if '__file__' in locals() else os.getcwd()
    
    # 어휘 등급 목록이 포함된 엑셀 파일 및 시트 이름
    WORD_LIST_EXCEL_FILE = "붙임1_★최종 공개용_국어 기초 어휘 선정 및 어휘 등급화 목록 전체.xlsx"
    WORD_LIST_SHEET_NAME = "전체(1~5등급), 40,000개"

    # 분석할 텍스트가 담긴 입력 파일 이름
    INPUT_TEXT_FILE = "지문모음.txt"
    # 분석 결과를 저장할 출력 엑셀 파일 이름
    OUTPUT_EXCEL_FILE = "ERI_지문별_최종분석결과.xlsx"

    # --- ERI 계산 계수 및 샘플링 설정 ---
    # ERI 양적 평가(eri_q) 계산에 사용되는 회귀식의 각 변수별 계수
    COEFFICIENTS = {'X2': -0.060, 'Z': 0.145, 'K_avg': 0.110, 'K_avg_Z': 0.024, 'intercept': 9.075}
    # 지문이 길 경우, 분석에 사용할 표본 텍스트의 어절 수
    SAMPLE_WORD_COUNT = 100
    # 표본 추출 시 지문의 (시작, 중간, 끝) 부분에서 가져올 어절의 비율
    SAMPLE_SPLIT_RATIO = (30, 40, 30)

    # --- 형태소 분석 및 문장 복잡도 계산 규칙 ---
    # 표제어(lemma)를 추출할 때 고려할 품사(POS) 태그 목록
    ALLOWED_POS_FOR_LEMMA = {"NNG", "VV", "VA", "MAG", "IC", "MM"} # 일반명사, 동사, 형용사, 일반부사, 감탄사, 관형사
    # 문장 분리를 위한 정규 표현식 (마침표, 물음표, 느낌표 기준)
    SENTENCE_SPLIT_REGEX = r'[.?!]\s*'
    # 절(clause)을 나누는 기준으로 사용될 연결 어미
    CLAUSE_CONNECTORS = {'고', '면', '서', '지만', '으며', '으면서', '자', '니까', '거나', '하며', '도', '과'}
    # 절(clause)을 나누는 기준으로 사용될 접속 부사
    CLAUSE_CONJUNCTIONS = {'그리고', '또는', '하지만', '그래서', '그럼에도'}
    # 서술어(predicate)로 간주할 품사 태그 목록
    PREDICATE_POS = {'VV', 'VA', 'XR', 'XSV', 'VXV', 'VXA', 'VX', 'VCP', 'VCN'}
    # 수식어(modifier)로 간주할 품사 태그 목록
    MODIFIER_POS = {'MM', 'MAG', 'MAJ', 'IC'} # 관형사, 일반부사, 접속부사, 감탄사
    # 명사구(noun phrase)로 판단할 품사 태그 목록 (서술어가 없는 절을 판단할 때 사용)
    NOUN_PHRASE_POS = {'NNG','NNP','NP','JKS','JKO','JKB','JKG','JKC','JKV','JKQ','JX','JC', 'SP','SY','MM','MAJ','MAG','IC'}
    # 명사구의 일부로 취급할 특정 형태소
    NOUN_PHRASE_SURFACE = {'등', '등을', '등이', '등과', '등도', '등만'}

# ==============================================================================
# 2. 데이터 로딩 및 분석 클래스
# ==============================================================================
def load_word_grade_dictionary(config: ERIConfig) -> Dict[str, int]:
    """
    설정 파일에 명시된 엑셀 파일에서 어휘와 등급 정보를 불러와 딕셔너리 형태로 반환합니다.
    - Key: 어휘 (str)
    - Value: 등급 (int)
    """
    excel_path = os.path.join(config.BASE_DIR, config.WORD_LIST_EXCEL_FILE)
    print(f"어휘 등급 사전 로딩 중... (파일: {config.WORD_LIST_EXCEL_FILE})")
    try:
        df = pd.read_excel(excel_path, sheet_name=config.WORD_LIST_SHEET_NAME)
        
        # '등급' 열에서 숫자만 추출하는 함수 (예: "1등급" -> 1)
        def parse_grade(grade_str):
            match = re.match(r"(\d+)", str(grade_str))
            return int(match.group(1)) if match else None
            
        df['등급_숫자'] = df['등급'].apply(parse_grade)
        df.dropna(subset=['어휘', '등급_숫자'], inplace=True) # 어휘 또는 등급이 없는 행 제거
        word_grade_dict = {row['어휘']: int(row['등급_숫자']) for _, row in df.iterrows()}
        print(f"총 {len(word_grade_dict):,}개의 어휘를 사전에 추가했습니다.")
        return word_grade_dict
    except FileNotFoundError:
        print(f"오류: 어휘 목록 파일 '{excel_path}'를 찾을 수 없습니다.")
        return {}
    except Exception as e:
        print(f"오류: 엑셀 파일 처리 중 오류 발생: {e}")
        return {}

def load_passages_from_file(filepath: str) -> List[Tuple[str, str]]:
    """
    '지문명:내용' 형식으로 작성된 텍스트 파일에서 지문 목록을 불러옵니다.
    """
    doclist = []
    try:
        with open(filepath, "r", encoding="utf-8") as f:
            lines = [line.strip() for line in f if line.strip()] # 공백 라인 제외
    except FileNotFoundError:
        print(f"오류: 입력 파일 '{filepath}'를 찾을 수 없습니다.")
        return []
        
    current_title, current_content = None, []
    # 파일을 한 줄씩 읽으며 '지문명:내용' 패턴을 찾아 분리
    for line in lines:
        match = re.match(r'^(.*?):(.*)$', line)
        if match:
            # 새로운 지문명이 나오면, 이전까지 저장된 지문명과 내용을 doclist에 추가
            if current_title is not None: 
                doclist.append((current_title, "\n".join(current_content)))
            current_title, content_part = match.groups()
            current_title = current_title.strip()
            current_content = [content_part.strip()]
        # ':' 패턴이 없는 라인은 현재 지문의 내용으로 간주
        elif current_title is not None:
            current_content.append(line)
            
    # 파일의 마지막 지문을 doclist에 추가
    if current_title is not None: 
        doclist.append((current_title, "\n".join(current_content)))
        
    return doclist


class ERICalculator:
    """ERI 지수 계산의 핵심 로직을 담당하는 클래스"""
    def __init__(self, config: ERIConfig, word_grade: Dict[str, int]):
        self.config = config
        self.word_grade = word_grade
        self.kiwi = Kiwi() # 형태소 분석기 초기화

    def _tokenize_to_morphs(self, text: str) -> List[Tuple[str, str]]:
        """텍스트를 형태소 분석하여 (형태, 품사)의 리스트로 반환하는 내부 함수"""
        return [(token.form, token.tag) for token in self.kiwi.tokenize(text)]

    def get_sample_text(self, full_text: str) -> str:
        """
        전체 텍스트에서 분석에 사용할 표본 텍스트를 추출합니다.
        지문이 설정된 샘플 크기(100어절)보다 길 경우, 시작/중간/끝 부분에서 지정된 비율로 텍스트를 가져옵니다.
        """
        words = full_text.split()
        total_words = len(words)
        s1, s2, s3 = self.config.SAMPLE_SPLIT_RATIO
        
        if total_words <= self.config.SAMPLE_WORD_COUNT: 
            return full_text # 100어절 이하면 전체 텍스트 사용
            
        part1 = words[:s1]
        mid_start = max(s1, (total_words - s2) // 2) # 중간 부분 시작점 계산
        part2 = words[mid_start : mid_start + s2]
        part3 = words[-s3:]
        
        return " ".join(part1 + part2 + part3)

    def extract_lemmas(self, text: str) -> Set[str]:
        """
        텍스트에서 표제어(기본형 단어)를 추출합니다.
        - 특정 품사(명사, 동사, 형용사 등)만 고려합니다.
        - '공부하다'와 같은 '명사+XSA' 형태의 복합어를 처리합니다.
        - 동사/형용사의 경우 '다'를 붙여 사전에 있는 형태와 일치시킵니다.
        """
        morphs = self.kiwi.analyze(text)[0][0]
        lemmas = set()
        i = 0
        while i < len(morphs):
            m = morphs[i]
            # '공부(NNG) + 하(XSA)' -> '공부하다' 와 같은 복합어 처리
            if (m.tag == 'NNG' and i + 1 < len(morphs) and morphs[i + 1].tag == 'XSA'):
                new_lemma = m.lemma + '하다'
                if new_lemma in self.word_grade: 
                    lemmas.add(new_lemma)
                i += 2; continue
            
            # 설정된 품사에 해당하면 표제어 추출
            if m.tag in self.config.ALLOWED_POS_FOR_LEMMA:
                # 한 글자짜리 특수문자 등은 제외
                if not (len(m.form) <= 1 and not m.form.isalpha()):
                    lemma = m.lemma
                    # 동사/형용사의 경우, '다'를 붙인 형태가 사전에 있다면 해당 형태로 추가
                    if m.tag in {"VA", "VV"} and (lemma + "다") in self.word_grade:
                        lemmas.add(lemma + "다")
                    else:
                        lemmas.add(lemma)
            i += 1
        return lemmas

    def _has_predicate(self, morphs: List[Tuple[str, str]]) -> bool:
        """형태소 리스트에 서술어 역할을 하는 품사가 있는지 확인"""
        return any(pos in self.config.PREDICATE_POS for _, pos in morphs)

    def _is_noun_phrase(self, morphs: List[Tuple[str, str]]) -> bool:
        """
        형태소 리스트가 명사구로만 이루어져 있는지 확인.
        'A, B, C 등' 과 같이 서술어 없이 명사만 나열된 절을 식별하기 위함.
        """
        cfg = self.config
        return all(pos in cfg.NOUN_PHRASE_POS or s in cfg.NOUN_PHRASE_SURFACE or s.startswith('등') for s, pos in morphs)

    def _smart_split_clauses(self, sentence: str) -> List[str]:
        """
        하나의 문장을 의미 단위의 절(clause)로 분리합니다.
        1. 연결어미, 접속부사, 쉼표 등을 기준으로 1차 분리합니다.
        2. 분리된 조각이 서술어 없는 명사구일 경우, 다음 조각과 합쳐 하나의 절로 만듭니다.
           (예: "사과, 배, 그리고 귤은 맛있다" -> ["사과, 배, 그리고 귤은", "맛있다"])
        """
        morphs = self._tokenize_to_morphs(sentence)
        temp_clauses, current_surfaces = [], []
        
        # 1. 연결어미, 접속부사, 쉼표 기준 1차 분리
        for surface, pos in morphs:
            current_surfaces.append(surface)
            is_connector = (pos == 'EC' and surface in self.config.CLAUSE_CONNECTORS) or \
                           (pos == 'MAJ' and surface in self.config.CLAUSE_CONJUNCTIONS) or \
                           (surface == ',' and pos in {'SP', 'SY'})
            if is_connector:
                temp_clauses.append("".join(current_surfaces).strip())
                current_surfaces = []
        if current_surfaces: 
            temp_clauses.append("".join(current_surfaces).strip())
            
        # 2. 서술어 없는 명사구를 후행 절과 병합
        merged, noun_phrase_buffer = [], []
        for clause in temp_clauses:
            clause_morphs = self._tokenize_to_morphs(clause)
            if not self._has_predicate(clause_morphs) and self._is_noun_phrase(clause_morphs):
                noun_phrase_buffer.append(clause.rstrip(",").strip())
            else:
                if noun_phrase_buffer:
                    merged.append(', '.join(noun_phrase_buffer + [clause.strip().lstrip(',')]))
                    noun_phrase_buffer = []
                else:
                    merged.append(clause)
        if noun_phrase_buffer: 
            merged.append(', '.join(noun_phrase_buffer))
            
        return [c for c in merged if c] # 비어있는 절 제거 후 반환

    def _calc_sentence_complexity(self, sentence: str) -> int:
        """
        문장 복잡도(K) 점수를 계산합니다. 점수가 높을수록 복잡한 문장입니다.
        점수 = Σ (절별 점수)
        절별 점수 = 기본 점수(문장 성분) + 수식 구조 점수 + 내포 구조 점수
        """
        clauses = self._smart_split_clauses(sentence)
        if not clauses: return 0
        
        total_score = 0
        for clause in clauses:
            morphs = self._tokenize_to_morphs(clause)
            
            # 문장 필수 성분(주어, 서술어, 목적어, 보어) 존재 여부 확인
            has_subj = any(p.startswith('N') for _, p in morphs)
            has_pred = self._has_predicate(morphs)
            has_obj = any(p == 'JKO' for _, p in morphs) # 목적격 조사
            has_comp = any(p == 'JKC' or (s == '이' and p == 'JKS' and i+1<len(morphs) and morphs[i+1][0]=='되다') for i,(s,p) in enumerate(morphs)) # 보격 조사, '되다/아니다' 앞의 주격조사
            
            # 1. 기본 점수 (문장 구조)
            base_score = 0
            if has_subj and has_pred:
                base_score = 1 # 주어+서술어
                if has_obj or has_comp: base_score = 2 # +목적어/보어
                if has_obj and has_comp: base_score = 3 # +목적어+보어
                
            # 2. 수식 구조 점수 (수식어 개수)
            mod_count = sum(1 for _,p in morphs if p in self.config.MODIFIER_POS) # 관형사, 부사 등
            mod_count += sum(1 for i in range(len(morphs)-1) if morphs[i][1]=='JKG' and morphs[i+1][1].startswith('NN')) # 관형격 조사 ('의')
            mod_count += sum(1 for i in range(len(morphs)-1) if morphs[i][1].startswith('NN') and morphs[i+1][1].startswith('NN')) # 복합 명사
            
            add1 = 0
            if 1 <= mod_count <= 3: add1 = 1
            elif 4 <= mod_count <= 6: add1 = 2
            elif mod_count > 6: add1 = 4
            
            # 3. 내포 구조 점수 (안은 문장, 인용 등)
            emb_count = sum(1 for _,p in morphs if p in {'ETM', 'ETN'}) # 전성어미 (관형사형, 명사형)
            emb_count += clause.count('"') + clause.count('“') + clause.count('”') # 인용 부호
            
            add2 = 0
            if 1 <= emb_count <= 5: add2 = emb_count * 3
            elif emb_count > 5: add2 = 18
            
            total_score += base_score + add1 + add2
            
        return total_score

    def calculate_eri(self, text: str) -> Dict:
        """
        주어진 텍스트에 대해 ERI 관련 지표들을 계산하고 결과를 딕셔너리로 반환합니다.
        - K_avg: 평균 문장 복잡도
        - X2: 쉬운 어휘(A등급, 1~3등급)의 개수
        - Z: 어려운 어휘(C등급 외, 등급 사전에 없는)의 개수
        - ERI_양적평가: 위 변수들을 회귀식에 대입한 값
        """
        # 1. 표본 텍스트 추출 및 문장 분리
        sample_text = self.get_sample_text(text)
        sentences = [s.strip() for s in re.split(self.config.SENTENCE_SPLIT_REGEX, sample_text) if s.strip()]
        
        # 2. K_avg (평균 문장 복잡도) 계산
        k_avg = (sum(self._calc_sentence_complexity(s) for s in sentences) / len(sentences)) if sentences else 0
        
        # 3. 표제어 추출 및 등급별 분류
        lemmas = self.extract_lemmas(sample_text)
        a_set = {lemma for lemma in lemmas if self.word_grade.get(lemma, 0) in {1, 2, 3}} # 1~3등급 어휘
        outside_set = {lemma for lemma in lemmas if lemma not in self.word_grade} # 사전에 없는 어휘
        
        x2, z = len(a_set), len(outside_set)
        
        # 4. ERI 양적 평가 점수 계산 (회귀식 적용)
        c = self.config.COEFFICIENTS
        eri_q = (c['X2']*x2) + (c['Z']*z) + (c['K_avg']*k_avg) + (c['K_avg_Z']*(k_avg*z)) + c['intercept']
        
        return {
            "표본텍스트": sample_text, "문장수": len(sentences),
            "K_avg(문장복잡도)": round(k_avg, 2), "X2(A등급 어휘 개수)": x2,
            "Z(C등급 외 어휘 개수)": z, "A등급 어휘": ", ".join(sorted(list(a_set))),
            "C등급 외 어휘": ", ".join(sorted(list(outside_set))),
            "ERI_양적평가": round(eri_q, 1),
        }

# ==============================================================================
# 3. GUI 및 결과 저장
# ==============================================================================
def qualitative_input_gui(eri_results):
    """
    Tkinter를 사용하여 ERI 질적 평가 점수를 입력받는 GUI 창을 생성합니다.
    - 스크롤 기능을 통해 많은 지문도 입력 가능하도록 처리합니다.
    - Tab, Shift+Tab, 위/아래 화살표 키로 입력 필드 간 이동이 가능합니다.
    """
    def save_scores():
        """'확인' 버튼 클릭 시, 입력된 점수를 eri_results에 저장하고 최종 지수를 계산"""
        for idx, var in enumerate(entry_vars):
            name = eri_results[idx]["지문명"]
            score_str = var.get().strip()
            try:
                score = float(score_str) if score_str else 0.0
                if not -3 <= score <= 3: raise ValueError
            except ValueError:
                messagebox.showerror("입력 오류", f"'{name}'의 점수는 -3에서 3 사이의 숫자로 입력하세요.")
                entries[idx].focus_set()
                return # 유효하지 않은 값이 있으면 저장 중단
            
            eri_results[idx]['ERI_질적평가'] = score
            eri_results[idx]['ERI_최종지수'] = round(eri_results[idx]['ERI_양적평가'] + score, 1)
        root.destroy()

    def cancel():
        """'취소' 버튼 또는 창 닫기 버튼 클릭 시, 확인 메시지를 띄우고 프로그램을 종료"""
        if messagebox.askokcancel("취소", "계산을 취소하시겠습니까? 결과가 저장되지 않습니다."):
            for r in eri_results:
                r['ERI_질적평가'] = '취소'
                r['ERI_최종지수'] = '취소'
            root.destroy()
            
    # --- GUI 창 기본 설정 ---
    root = tk.Tk()
    root.title("질적평가 점수 입력")
    root.attributes("-topmost", True) # 창을 항상 위로
    root.grab_set() # 다른 창과 상호작용 방지
    root.focus_force()

    # --- 스크롤 가능한 프레임 생성 ---
    outer = tk.Frame(root, bg="#fafbfc")
    outer.pack(fill="both", expand=True)

    canvas = tk.Canvas(outer, borderwidth=0, highlightthickness=0, bg="#fafbfc")
    frame = tk.Frame(canvas, bg="#fafbfc") # 실제 위젯이 담길 프레임
    vscroll = tk.Scrollbar(outer, orient="vertical", command=canvas.yview)
    canvas.configure(yscrollcommand=vscroll.set)
    vscroll.pack(side="right", fill="y")
    canvas.pack(side="left", fill="both", expand=True)
    canvas.create_window((0, 0), window=frame, anchor="nw")

    def on_configure(event): canvas.configure(scrollregion=canvas.bbox("all")) # 프레임 크기 변경 시 스크롤 영역 재설정
    frame.bind("<Configure>", on_configure)

    def _on_mousewheel(event): canvas.yview_scroll(int(-1*(event.delta/120)), "units") # 마우스 휠 스크롤
    canvas.bind_all("<MouseWheel>", _on_mousewheel)

    # --- 키보드 네비게이션 및 자동 스크롤 함수 ---
    def move_focus(idx, delta):
        """
        입력 필드(Entry) 사이를 위/아래 화살표나 탭 키로 이동시킵니다.
        포커스가 이동할 때, 해당 위젯이 보이도록 뷰를 자동으로 스크롤합니다.
        """
        n = len(entries)
        next_idx = (idx + delta + n) % n # 다음 인덱스 계산 (순환 구조)
        entries[next_idx].focus_set()
        
        # UI가 업데이트되어야 정확한 위젯 위치를 알 수 있음
        root.update_idletasks() 

        # 다음 위젯의 y 좌표와 높이
        entry_y = entries[next_idx].winfo_y()
        entry_height = entries[next_idx].winfo_height()
        
        # 현재 캔버스 뷰의 y 좌표와 높이
        canvas_top_y = canvas.canvasy(0)
        canvas_height = canvas.winfo_height()

        # 위젯이 뷰의 위쪽 경계보다 위에 있을 경우
        if entry_y < canvas_top_y:
            canvas.yview_moveto(entry_y / frame.winfo_height())
        # 위젯이 뷰의 아래쪽 경계보다 아래에 있을 경우
        elif entry_y + entry_height > canvas_top_y + canvas_height:
            canvas.yview_moveto((entry_y + entry_height - canvas_height) / frame.winfo_height())
            
        return "break" # 이벤트 전파 중단
    
    # --- GUI 위젯 배치 ---
    label_font = ("맑은 고딕", 15, "bold")
    input_font = ("맑은 고딕", 13)
    head_bg = "#e5eefc"

    tk.Label(frame, text="지문명", font=label_font, bg=head_bg, width=18, anchor="center", pady=9).grid(row=0, column=0, padx=(26,10), pady=10, sticky="ew")
    tk.Label(frame, text="질적 평가 점수\n(-3~3)", font=label_font, bg=head_bg, width=16, anchor="center", pady=9, justify="center").grid(row=0, column=1, padx=(10,16), pady=10, sticky="ew")

    entry_vars, entries = [], []
    for i, result in enumerate(eri_results):
        # 지문명 표시 (Text 위젯 사용으로 자동 줄바꿈 지원)
        text_widget = tk.Text(frame, height=2, width=26, font=input_font, wrap='word', bg="#f7fafc", relief="groove", borderwidth=2)
        text_widget.insert("1.0", result['지문명'])
        text_widget.config(state="disabled") # 편집 불가
        text_widget.grid(row=i+1, column=0, padx=(26,10), pady=7, sticky="ew")
        
        # 점수 입력 필드
        var = tk.StringVar()
        entry = tk.Entry(frame, textvariable=var, font=input_font, justify="center", width=13, relief="groove")
        entry.grid(row=i+1, column=1, padx=(10,16), pady=7, sticky="ew")
        
        # 키 바인딩 설정
        entry.bind("<Tab>", lambda event, idx=i: move_focus(idx, 1))
        entry.bind("<Shift-Tab>", lambda event, idx=i: move_focus(idx, -1))
        entry.bind("<Down>", lambda event, idx=i: move_focus(idx, 1))
        entry.bind("<Up>", lambda event, idx=i: move_focus(idx, -1))
        
        entry_vars.append(var)
        entries.append(entry)
        
        if i == 0: root.after(150, entry.focus_set) # 창이 열리면 첫 번째 입력 필드에 자동 포커스
    
    # 확인/취소 버튼
    btn_outer = tk.Frame(outer, bg="#fafbfc")
    btn_outer.pack(side="right", fill="y", padx=(8,32), pady=45)
    btn_outer.grid_rowconfigure(0, weight=1); btn_outer.grid_rowconfigure(1, weight=1)
    tk.Button(btn_outer, text="확인", width=8, height=2, font=input_font, command=save_scores, bg="#0077d9", fg="white").grid(row=0, column=0, pady=(0,18), sticky="ew")
    tk.Button(btn_outer, text="취소", width=8, height=2, font=input_font, command=cancel).grid(row=1, column=0, pady=(0,0), sticky="ew")

    # 창 크기 및 위치 조절
    root.update_idletasks() # 위젯 크기 계산을 위해 UI 업데이트
    width, height = 740, min(540, 85 + 49 * max(3, len(eri_results))) # 지문 수에 따라 창 높이 조절
    x, y = (root.winfo_screenwidth()//2) - (width//2), (root.winfo_screenheight()//2) - (height//2) # 화면 중앙에 위치
    root.geometry(f"{width}x{height}+{x}+{y}")
    root.deiconify()
    root.lift()
    root.focus_force()
    root.protocol("WM_DELETE_WINDOW", cancel) # 창 닫기 버튼을 '취소'와 동일하게 처리
    root.mainloop()

# ==============================================================================
# 4. 메인 실행 로직
# ==============================================================================
def main():
    """프로그램의 주 실행 흐름을 제어"""
    # 1. 설정 및 필요 데이터 로딩
    config = ERIConfig()
    word_grade_dict = load_word_grade_dictionary(config)
    if not word_grade_dict:
        print("어휘 사전을 불러오지 못해 프로그램을 종료합니다."); return

    passage_path = os.path.join(config.BASE_DIR, config.INPUT_TEXT_FILE)
    passages = load_passages_from_file(passage_path)
    if not passages:
        print("분석할 지문이 없어 프로그램을 종료합니다."); return

    # 2. 각 지문에 대해 ERI 양적 평가 실행
    calculator = ERICalculator(config, word_grade_dict)
    all_results = []
    print("\nERI 양적 평가를 시작합니다...")
    for title, content in passages:
        try:
            result = {"지문명": title, **calculator.calculate_eri(content)}
            all_results.append(result)
            print(f"- '{title}' 분석 완료 (양적 ERI: {result['ERI_양적평가']})")
        except Exception as e:
            print(f"- '{title}' 분석 중 오류 발생: {e}")

    if not all_results:
        print("모든 지문 분석에 실패하여 프로그램을 종료합니다."); return

    # 3. GUI를 통해 질적 평가 점수 입력받기
    qualitative_input_gui(all_results)
    
    # 4. 사용자가 취소했는지 확인 후 결과 처리
    if all_results and all_results[0].get('ERI_최종지수') == '취소':
        print("\n사용자가 질적 평가를 취소했습니다. 프로그램을 종료합니다.")
        return

    # 5. 최종 결과를 엑셀 파일로 저장
    outpath = os.path.join(config.BASE_DIR, config.OUTPUT_EXCEL_FILE)
    # 엑셀에 저장할 컬럼 순서 지정
    desired_cols = ["지문명", "문장수", "표본텍스트", "K_avg(문장복잡도)", "X2(A등급 어휘 개수)",
                    "Z(C등급 외 어휘 개수)", "A등급 어휘", "C등급 외 어휘", "ERI_양적평가", "ERI_질적평가", "ERI_최종지수"]
    df = pd.DataFrame(all_results)
    df = df.reindex(columns=desired_cols)
    df.to_excel(outpath, index=False)

    # 6. 저장된 엑셀 파일 열어서 서식 꾸미기
    wb = load_workbook(outpath)
    ws = wb.active
    # 컬럼 너비 설정
    col_widths = {"A": 25, "B": 10, "C": 48, "D": 19, "E": 19, "F": 19, "G": 36, "H": 36, "I": 13, "J": 13, "K": 13}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width
    
    # 셀 정렬 및 텍스트 줄바꿈 설정
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            col_letter = get_column_letter(cell.column)
            if col_letter == 'A': # 지문명
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False, shrink_to_fit=True)
            elif col_letter == 'C': # 표본 텍스트
                 cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            elif col_letter in ["G", "H"]: # 어휘 목록
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            else: # 나머지
                cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 'ERI_최종지수' 컬럼 폰트를 굵은 빨간색으로 강조
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=11, max_col=11):
        for cell in row:
            cell.font = Font(bold=True, color="FF0000")
            
    wb.save(outpath)

    # 7. 최종 결과 콘솔에 출력
    print("\n" + "="*50)
    for r in all_results:
        print(f"'{r['지문명']}' | ERI 최종지수: {r['ERI_최종지수']}")
    print("="*50)
    print(f"\n전체 지문 분석결과 저장 완료! -> {outpath}")


if __name__ == '__main__':
    main()